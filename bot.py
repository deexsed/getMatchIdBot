import json
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackContext, CallbackQueryHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import telegram.ext.filters as filters
from config import TG_BOT_TOKEN
from datetime import datetime

# Функция для загрузки списка героев из JSON-файла
def load_heroes(filename='heroes.json'):
    with open(filename, 'r', encoding='utf-8') as file:
        data = json.load(file)
        return data['heroes']

# Загружаем список героев при запуске бота
HEROES = load_heroes()

# Создаем новый Excel файл и добавляем заголовки, если он не существует
def create_excel_file():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Match Data"
    sheet.append(["Date", "Nickname", "Match ID", "Hero", "Outcome"])  # Добавляем столбец "Дата"
    workbook.save("matchStat.xlsx")

# Обработчик команды /start для бота
async def start(update: Update, context: CallbackContext) -> None:
    await update.message.reply_text('Введите ID матча:')
    context.user_data['waiting_for_match_id'] = True

# Обработчик сообщений для обработки ввода пользователя
async def handle_message(update: Update, context: CallbackContext) -> None:
    if context.user_data.get('waiting_for_match_id'):
        nickname = f"@{update.message.from_user.username}" if update.message.from_user.username else update.message.from_user.first_name
        
        match_id = update.message.text.strip()

        if match_id.isdigit() or (len(match_id) >= 8 or len(match_id) <= 16):
            context.user_data['match_id'] = match_id
            await send_hero_selection(update)  # Отправляем клавиатуру с героями
            context.user_data['waiting_for_hero'] = True
            context.user_data['nickname'] = nickname
            context.user_data['waiting_for_match_id'] = False
        else:
            await update.message.reply_text('Пожалуйста, введите корректный ID матча (только цифры, длина от 8 до 16 символов включительно).')

async def send_hero_selection(update: Update) -> None:
    # Разбиваем список героев на строки по 2 кнопки
    keyboard = []
    for i in range(0, len(HEROES), 2):
        row = [InlineKeyboardButton(hero, callback_data=hero) for hero in HEROES[i:i + 2]]
        keyboard.append(row)
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text('Выберите имя героя, на котором вы играли:', reply_markup=reply_markup)

async def send_match_outcome_selection(message) -> None:
    keyboard = [
        [InlineKeyboardButton("Победа", callback_data='win')],
        [InlineKeyboardButton("Поражение", callback_data='lose')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await message.reply_text('Выберите исход матча:', reply_markup=reply_markup)

async def handle_hero_selection(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Подтверждаем нажатие кнопки

    # Проверяем, что callback_data соответствует герою
    if query.data not in HEROES:
        return  # Игнорируем, если это не выбор героя

    # Проверяем наличие необходимых данных
    match_id = context.user_data.get('match_id')
    nickname = context.user_data.get('nickname')

    if match_id is None or nickname is None:
        # Создаем новый объект Update для передачи в функцию start
        new_update = Update(update_id=query.id, message=query.message)
        
        await start(new_update, context)  # Запускаем процесс записи заново
        return

    hero_name = query.data  # Получаем имя героя из callback_data
    context.user_data['hero_name'] = hero_name

    await send_match_outcome_selection(query.message)  # Отправляем кнопки для выбора исхода матча


async def handle_match_outcome(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Подтверждаем нажатие кнопки

    # Проверяем, что callback_data соответствует исходу матча
    if query.data not in ['win', 'lose']:
        return  # Игнорируем, если это не выбор исхода

    match_outcome = query.data  # Получаем исход матча из callback_data
    match_id = context.user_data.get('match_id')
    nickname = context.user_data.get('nickname')
    hero_name = context.user_data.get('hero_name')

    if match_id is None or nickname is None or hero_name is None:
        # Создаем новый объект Update для передачи в функцию start
        new_update = Update(update_id=query.id, message=query.message)
        
        await start(new_update, context)  # Запускаем процесс записи заново
        return

    save_data(match_id, nickname, hero_name, match_outcome)

    await query.message.reply_text('Данные сохранены!')
    await send_restart_button(query.message)  # Отправляем кнопку для начала нового ввода
    
    context.user_data.clear()

async def send_restart_button(update: Update) -> None:
    keyboard = [
        [InlineKeyboardButton("Записать новый матч", callback_data='restart')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.reply_text('Хотите записать новый матч?', reply_markup=reply_markup)

async def handle_restart(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Подтверждаем нажатие кнопки

    # Создаем новый объект Update для передачи в функцию start
    new_update = Update(update_id=query.id, message=query.message)
    
    await start(new_update, context)  # Запускаем процесс записи заново

def save_data(match_id, nickname, hero_name, match_outcome):
    try:
        workbook = load_workbook("matchStat.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        create_excel_file()
        workbook = load_workbook("matchStat.xlsx")
        sheet = workbook.active

    # Если это новая таблица, задаем стили для заголовков
    if sheet.max_row == 1:
        header_font = Font(bold=True, color="FFFFFF")  # Белый шрифт
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синий фон
        for cell in sheet[1]:  # Заголовки находятся в первой строке
            cell.font = header_font
            cell.fill = header_fill

    # Получаем текущую дату и время
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Добавляем новую строку с датой
    new_row = [current_time, nickname, match_id, hero_name, match_outcome]
    sheet.append(new_row)

    # Определяем цвет для текущей строки
    current_row = sheet.max_row
    if current_row % 2 == 0:  # Четные строки
        fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый фон
    else:  # Нечетные строки
        fill_color = PatternFill(start_color="e6f3fb", end_color="e6f3fb", fill_type="solid")  # Полупрозрачный синий

    # Применяем цвет к ячейкам новой строки
    for col in range(1, 6):  # Теперь у нас 5 колонок
        cell = sheet.cell(row=current_row, column=col)
        cell.fill = fill_color
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

    # Изменяем цвет шрифта в зависимости от исхода матча
    outcome_cell = sheet.cell(row=current_row, column=5)  # Столбец "Outcome" теперь 5-й
    if match_outcome == 'win':
        outcome_cell.font = Font(color="00FF00")  # Зеленый цвет для победы
    else:
        outcome_cell.font = Font(color="FF0000")  # Красный цвет для поражения

    workbook.save("matchStat.xlsx")

    # Выводим информацию о добавленной записи в консоль
    print(f"{current_time} - Добавлена запись: {nickname}, {match_id}, {hero_name}, {match_outcome}")

def main() -> None:
    print("Бот успешно запущен!\n")
    application = ApplicationBuilder().token(TG_BOT_TOKEN).build()

    application.add_handler(CommandHandler('start', start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Обработчик для кнопок выбора героев
    application.add_handler(CallbackQueryHandler(handle_hero_selection, pattern='^(' + '|'.join(HEROES) + ')$'))
    
    # Обработчик для кнопок выбора исхода матча
    application.add_handler(CallbackQueryHandler(handle_match_outcome, pattern='^(win|lose)$'))
    
    # Обработчик для кнопки перезапуска
    application.add_handler(CallbackQueryHandler(handle_restart, pattern='^restart$'))

    application.run_polling()

if __name__ == '__main__':
    main()