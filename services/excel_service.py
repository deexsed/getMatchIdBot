"""
Сервис для работы с Excel файлами
"""
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import time
from utils.constants import (
    EXCEL_FILENAME, WIN_COLOR, LOSE_COLOR, HEADER_BG_COLOR,
    HEADER_FONT_COLOR, ODD_ROW_COLOR, EVEN_ROW_COLOR
)

logger = logging.getLogger(__name__)

def create_excel_file():
    """Создает новый Excel файл с заголовками"""
    workbook = Workbook()
    
    # Основной лист с данными матчей
    match_sheet = workbook.active
    match_sheet.title = "Match Data"
    
    # Основные заголовки
    headers = [
        "Date", 
        "Telegram Nickname",
        "Match ID", 
        "Hero",
        "Outcome",
        "Hero Winrate"
    ]
    
    # Записываем основные заголовки
    for col, header in enumerate(headers, start=1):
        match_sheet.cell(row=1, column=col, value=header)
    
    # Форматирование заголовков
    for col in range(1, 7):
        cell = match_sheet.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Устанавливаем ширину колонок
    for col in range(1, 7):
        match_sheet.column_dimensions[get_column_letter(col)].width = 15

    # Создаем лист статистики
    stats_sheet = workbook.create_sheet(title="Statistics")
    
    # Заголовки для общей статистики
    stats_sheet['A1'] = "Общая статистика"
    stats_sheet['A2'] = "Всего игр"
    stats_sheet['A3'] = "Общий винрейт"
    
    # Заголовки для статистики по героям
    stats_sheet['A6'] = "Статистика по героям"
    stats_sheet['A7'] = "Герой"
    stats_sheet['B7'] = "Игр"
    stats_sheet['C7'] = "Винрейт"

    # Форматирование заголовков
    for cell in [stats_sheet['A1'], stats_sheet['A6']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in [stats_sheet['A7'], stats_sheet['B7'], stats_sheet['C7']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F3FB", end_color="E6F3FB", fill_type="solid")

    # Устанавливаем ширину колонок
    for col in ['A', 'B', 'C']:
        stats_sheet.column_dimensions[col].width = 15
    
    workbook.save(EXCEL_FILENAME)

def update_stats_section(sheet, nickname=None):
    """Обновляет секцию статистики в таблице"""
    # ... (оставшийся код функции без изменений)

def save_match_to_excel(match_id, tg_nickname, hero_name, match_outcome):
    """Сохраняет данные матча в Excel"""
    max_attempts = 3
    attempt = 0
    
    while attempt < max_attempts:
        try:
            # ... (код сохранения в Excel)
            pass
        except Exception as e:
            logger.error(f"Ошибка при сохранении данных: {e}")
            raise 